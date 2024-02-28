import os
import json
import requests
import openpyxl
from datetime import datetime
import csv
import pandas as pd
from dotenv import load_dotenv
from pathlib import Path

#load environment
dotenv_path = Path('.env')
load_dotenv(dotenv_path=dotenv_path)

input_file_path = "./COMBINE_Cohort_Inclusion.xlsx"
lab_file_path = "./Laborwerte_mit_LOINC_300.xlsx"
input_patient_map = "./combine_pat_ids_PseudonymisiertgPAS.csv"

fhir = "https://URL_TO_FHIR_SERVER"
username = os.environ.get('FHIR_USERNAME')
password = os.environ.get('FHIR_PASSWORD')

# fhir_dst = ""
# username_dst = os.environ.get('FHIR_DST_USERNAME')
# password_dst = os.environ.get('FHIR_DST_USERNAME')
# headers = {"Content-Type": "application/fhir+json", 'j_username': username_dst, 'j_password': password_dst}

#column of patient IDs
PAT_COL = 'A'
#Number of FHIR resources per http GET request
MAX_COUNT = 100
#Save each time period per patientrow in an individual json, default false
SINGLE_PERIODS = False
#Replace DIZ-pseudonym in FHIR resources with patient id
REPLACE_PAT_ID = False
#Naming systems
pat_system = "https://URL_TO_FHIR_SERVER/identifiers/patient-id"
loinc_system = "http://loinc.org"

#Process get request to DIZ-FHIR Server
def send_request(in_query):
    url = fhir + "/" + in_query
    response = session.get(url)
    # print(url)
    try:
        content = json.loads(response.content.decode('utf-8'))
    except json.decoder.JSONDecodeError:
        print(f"code = {response.status_code}")
        print(f"content= {response.content}")
        return None, None
    
    return response, content

#Turns an integer index into its Excel representation
def i_to_col(index: int) -> str:
    result = ""
    while index >= 0:
        result = chr(index % 26 + 65) + result
        index = (index // 26) - 1
    return result

#Return lab codes from lab_file_path table
def get_lab_codes() -> list:
    start_row = 2
    col = 'C'
    sheet = openpyxl.load_workbook(lab_file_path, data_only=True).active
    labcodes = []
    
    while(True):
        val = sheet[col+str(start_row)].value
        if val is not None:
            labcodes.append(str(val))
        if val is None and sheet['A'+str(start_row)].value is None:
            break
        start_row += 1

    print(labcodes)
    labcodes = ",".join([f"{loinc_system}|" + s for s in labcodes])
    return labcodes

#Get PatientIDs and write them to .csv to supply it to your THS for pseudonimization
def get_patIDs():
    workbook = openpyxl.load_workbook(input_file_path, data_only=True)
    sheets = workbook.sheetnames
    pat_ids = set()

    for s in sheets:
        ws = workbook[s]
        _,start_index = get_row_index(worksheet=ws)
        i = start_index
        while(True):
            val = ws['A'+str(i)].value
            if val is not None:
                #val = str(val)[:-1]
                val = str(val)
                pat_ids.add(val)
            else:
                break
            i+=1

    cw = csv.writer(open('combine_pat_ids_.csv', 'w', newline=''))
    for elem in pat_ids:
        cw.writerow([str(elem).strip()])

def get_patient_map():
    return pd.read_csv(input_patient_map, header=0, index_col=0, sep=";").squeeze("columns").to_dict()

def get_message(input_response):
    return json.loads(input_response.content.decode('utf-8'))["issue"][0]["details"]["text"]

def get_row_index(worksheet):
    for row in range(1, worksheet.max_row, 1):
        if worksheet.cell(row, column=1).value == 'PatientID':
            header_index = row
            pat_start_index = row+1
            break
    return header_index, pat_start_index

if __name__ == "__main__":
    # get_patIDs()
    # quit()

    if not os.path.exists(f"output"):
        os.makedirs(f"output/json")
        os.makedirs(f"output/csv")
    
    # src session
    session = requests.Session()
    session.auth = (username, password)
    session.post(fhir)
    print(f"Connected to {fhir}")

    # # dst session
    # session_dst = requests.Session()
    # session_dst.auth = (username_dst, password_dst)
    # session_dst.post(fhir_dst)
    # print(f"Connected to destination {fhir_dst}")
    
    patient_map = get_patient_map()
    lab_codes = get_lab_codes()
    workbook = openpyxl.load_workbook(input_file_path, data_only=True)
    sheets = workbook.sheetnames

    #Iterate through cohort sheets
    for s in sheets:

        output_csv = []
        periods = []
        header_dict = {}
        ws = workbook[s]
        print(s)
        
        i = 0
        header_index, pat_start_index = get_row_index(worksheet=ws)

        while(True):
            try:
                val = ws[i_to_col(i)+str(header_index)].value.replace(' ','')
            except AttributeError as attr:
                break

            if("von" in val):
                idx = val.split('(')[1].split(')')[0]
                periods.append(idx)
                header_dict[idx] = [i_to_col(i), i_to_col(i+1), i_to_col(i+2)]
                i += 3
                continue
            i += 1
        print(header_dict)

        #Iterate through patient rows
        for j in range(pat_start_index, ws.max_row+1, 1):
            pat_id = ws[PAT_COL+str(j)].value
            try:
                # pseud_id = patient_map[int(str(pat_id)[:-1])]
                pseud_id = patient_map[int(str(pat_id))]
            except KeyError as keyerror:
                print(f"Patient not found: {keyerror}")
                continue

            #extend existing file if it exists else start from scratch
            if os.path.exists(f"./output/{s}_{pseud_id}.json"):
                json_data = open(f"./output/{s}_{pseud_id}.json")
                d = json.load(json_data)
                json_data.close()

                print(f"Appending file ./output/{s}_{pseud_id}.json")
                bundle = d
                entry = d["entry"]
            else:
                bundle = {
                    "resourceType": "Bundle",
                    "id": pseud_id,
                    "type": "transaction",
                    "entry": []
                }
                entry = []

            #Iterate through cohort time periods (Pre, Initial, Peak, FollowUp)
            for p in periods:        
                start,ideal,ende = [x+str(j) for x in header_dict[p]]
                start_val = str(ws[start].value).split(' ')[0]
                ideal_val = str(ws[ideal].value).split(' ')[0]
                ende_val = str(ws[ende].value).split(' ')[0]

                if(start_val is None):
                    break
                
                #Continue if no correct time value in row
                try:
                    isinstance(datetime.strptime(start_val, "%Y-%m-%d"),datetime)
                    isinstance(datetime.strptime(ideal_val, "%Y-%m-%d"),datetime)
                    isinstance(datetime.strptime(ende_val, "%Y-%m-%d"),datetime)
                except ValueError:
                    continue

                print(s, p, pseud_id, start_val, ideal_val, ende_val)
                
                page = 1

                response,content = send_request(f"Observation?code={lab_codes}&patient.identifier={pseud_id}&date=gt{start_val}&date=lt{ende_val}&_count={MAX_COUNT}&_page={page}")
                if response is None:
                    continue
                elif response.status_code > 400:
                    print(f"{response.status_code}: {get_message(response)}")
                    continue
                
                try:
                    if content["total"] == 0:
                        print(f"No lab data found!")
                        continue
                    elif content["total"] > MAX_COUNT:
                        entry_tmp = content["entry"]
                        cont = content
                        while(cont):
                            page += 1
                            resp,cont = send_request(f"Observation?code={lab_codes}&patient.identifier={pseud_id}&date=gt{start_val}&date=lt{ende_val}&_count={MAX_COUNT}&_page={page}")
                            if response.status_code > 400:
                                print(f"{response.status_code}: {get_message(response)}")
                            
                            if "entry" in cont:
                                entry_tmp += cont["entry"]
                            else:
                                content["entry"] = entry_tmp
                                break
                except KeyError as error:
                    print(f"{pseud_id}/{pat_id}: {error}")

                for e in content["entry"]:
                    if REPLACE_PAT_ID:
                        e["resource"]["subject"]["identifier"]["system"] = pat_system
                        e["resource"]["subject"]["identifier"]["value"] = pat_id
                    del e["search"]
                    e["request"] = {
                        "method": "PUT",
                        "url": e["resource"]["resourceType"] + "/" + e["resource"]["id"]
                    }

                entry.extend(content["entry"])

                #save each period individually
                if(SINGLE_PERIODS):
                    #Adapt resource to make a transaction bundle
                    content["type"] = "transaction"
                    del content["link"]
                    del content["total"]
                    
                    with open(f"output/{s}_{str(j)}_{pseud_id}_{p}_{header_dict[p][1]}.json", "w") as f:
                        f.write(json.dumps(content, indent=4))

                try:
                    for item in entry:
                        value_set = {
                            "PID":pat_id,
                            "Date":item["resource"]["effectiveDateTime"],
                            "LOINC": "",
                            "Value":item["resource"]["valueQuantity"]["value"],
                            "Lower_bound": "",
                            "Upper_bound": "",
                            "Interpretation_Code": [],
                            "Interpretation_Display": []
                        }
                        for i in range(len(item["resource"]["code"]["coding"])):
                            if item["resource"]["code"]["coding"][i]["system"] == loinc_system:
                                value_set["LOINC"] = item["resource"]["code"]["coding"][i]["code"]
                                break
                        
                        if "referenceRange" in item["resource"]:
                            for i in range(len(item["resource"]["referenceRange"])):
                                if len(item["resource"]["referenceRange"]) > 1:
                                    print(item["resource"]["referenceRange"])
                                    quit()
                                if "low" in item["resource"]["referenceRange"][i]:
                                    value_set["Lower_bound"] = item["resource"]["referenceRange"][i]["low"]["value"]
                                if "high" in item["resource"]["referenceRange"][i]:
                                    value_set["Upper_bound"] = item["resource"]["referenceRange"][i]["high"]["value"]

                        interpret_code, interpret_display = [], []
                        if "interpretation" in item["resource"]:
                            for int in range(len(item["resource"]["interpretation"])):
                                for cod in range(len(item["resource"]["interpretation"][int]["coding"])):
                                    interpret_code.append(item["resource"]["interpretation"][int]["coding"][cod]["code"])
                                    interpret_display.append(item["resource"]["interpretation"][int]["coding"][cod]["display"])

                        value_set["Interpretation_Code"] = interpret_code
                        value_set["Interpretation_Display"] = interpret_display
                        
                        # data = [{"PID":pat_id, "Date":item["resource"]["effectiveDateTime"], "LOINC":item["resource"]["code"]["coding"][0]["code"], "Value":item["resource"]["valueQuantity"]["value"]}]
                        output_csv.extend([value_set])
                except KeyError as keyerror:
                    print(keyerror)
                    pass

            if len(entry) > 0:
                bundle["entry"] = entry
                with open(f"output/json/{s}_{pseud_id}.json", "w") as f:
                    f.write(json.dumps(bundle, indent=4))

        with open(f"output/csv/{s}_output.csv", 'w', newline="") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=["PID", "Date", "LOINC", "Value", "Lower_bound", "Upper_bound", "Interpretation_Code", "Interpretation_Display"])
            writer.writeheader()
            writer.writerows(output_csv)

                #POST TO FHIR IF DESTINATION AVAILABLE
                # response = session_dst.post(fhir_dst, headers=headers, data=json.dumps(bundle))
                # if response.status_code >= 400:
                #     print(f"{response.status_code}: {get_message(response)}")